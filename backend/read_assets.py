#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
import json
from datetime import datetime
import sys

# Read Excel file
excel_file = r'c:\Apps\assethub-V2\AssetIT41.xlsx'

try:
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    
    print("=" * 80)
    print(f"📊 Asset File: {excel_file}")
    print("=" * 80)
    
    # Get headers
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)
    
    print(f"\n📋 Headers (Column {len(headers)}):")
    for i, h in enumerate(headers, 1):
        print(f"  {i:2d}. {h}")
    
    # Get data rows
    assets = []
    print(f"\n📦 Data Preview (First 10 rows):")
    print("-" * 80)
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        if row_idx > 11:  # Show only first 10 data rows
            break
        
        row_data = {}
        for col_idx, cell in enumerate(row):
            header = headers[col_idx] if col_idx < len(headers) else f"Col_{col_idx}"
            value = cell.value
            row_data[header] = value
        
        assets.append(row_data)
        
        print(f"\nRow {row_idx}:")
        for key, val in row_data.items():
            if val is not None:
                print(f"  {key:30s}: {val}")
    
    # Count total rows
    total_rows = ws.max_row - 1
    print(f"\n{'=' * 80}")
    print(f"✓ Total records: {total_rows}")
    print(f"{'=' * 80}")
    
    # Save to JSON for TypeScript processing
    json_output = {
        'headers': headers,
        'total_rows': total_rows,
        'sample_rows': assets
    }
    
    with open(r'c:\Apps\assethub-V2\backend\asset_data.json', 'w', encoding='utf-8') as f:
        json.dump(json_output, f, indent=2, ensure_ascii=False, default=str)
    
    print(f"\n✓ JSON saved to: c:\\Apps\\assethub-V2\\backend\\asset_data.json")
    
except Exception as e:
    print(f"❌ Error: {e}", file=sys.stderr)
    sys.exit(1)
